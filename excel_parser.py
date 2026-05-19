from pathlib import Path
import openpyxl
from datetime import datetime


def parse_banksalad_excel(file_path: str | Path) -> list[dict]:
    """뱅크샐러드 Excel 파일을 파싱하여 트랜잭션 목록 반환."""
    wb = openpyxl.load_workbook(file_path)

    # '가계부 내역' 시트 찾기
    sheet = None
    for name in wb.sheetnames:
        if "가계부" in name or "내역" in name:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    result = []

    for row in rows:
        if not row or row[0] is None:
            continue

        # 날짜
        date_val = row[0]
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]

        # 시간
        time_val = row[1]
        if hasattr(time_val, "strftime"):
            time_str = time_val.strftime("%H:%M")
        elif time_val is not None:
            time_str = str(time_val)[:5]
        else:
            time_str = ""

        # 금액
        try:
            amount = int(row[6]) if row[6] is not None else 0
        except (TypeError, ValueError):
            amount = 0

        result.append(
            {
                "date": date_str,
                "time": time_str,
                "type": str(row[2] or ""),
                "cat": str(row[3] or "미분류"),
                "subcat": str(row[4] or "미분류"),
                "desc": str(row[5] or ""),
                "amount": amount,
                "currency": str(row[7] or "KRW"),
                "method": str(row[8] or ""),
                "memo": str(row[9] or "") if len(row) > 9 else "",
            }
        )

    return result


def get_date_range(transactions: list[dict]) -> tuple[str, str]:
    dates = [t["date"] for t in transactions if t.get("date")]
    if not dates:
        return ("", "")
    return min(dates), max(dates)
